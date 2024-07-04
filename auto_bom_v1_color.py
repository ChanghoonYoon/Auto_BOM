import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# CSV 데이터 (수정된 버전)
csv_data = """
Category,Specification,Reference,Quantity
Capacitor,8pF,"C2,C3",2
Capacitor,20pF,C21,1
Capacitor,22pF,"C4,C5",2
Capacitor,24pF,"C29,C30,C31,C32,C33",5
Capacitor,27pF,C34,1
Capacitor,100nF,C20,1
Capacitor,0.1uF,"C1,C7,C8,C9,C11,C12,C13,C14,C15,C16,C17,C18,C26,C35,C39,C40",16
Capacitor,1uF,"C19,C25,C36",3
Capacitor,2.2uF,C10,1
Capacitor,4.7uF,"C37,C38",2
Capacitor,10uF,C6,1
Capacitor,10uF/16V/2012,C24,1
Capacitor,22uF/10V/2012,"C22,C23",2
Capacitor,47uF/10V/3216,"C27,C28",2
Resistor,0,"R15,R17,R55,R76,R83,R111,R112",7
Resistor,22,"R32,R33,R35,R38,R40,R42,R46,R50,R51,R52,R54,R56,R58,R60,R62,R64,R66,R68,R70,R71,R84,R85",22
Resistor,27R,"R72,R73",2
Resistor,100,"R20,R21,R22,R23,R24,R25,R86,R87,R88",9
Resistor,200,"R9,R10,R11,R12,R14,R39,R41",7
Resistor,1K,"R28,R36,R97,R98,R99,R100,R101",7
Resistor,1.5K,R30,1
Resistor,2K,"R8,R45,R53,R74,R79,R81",6
Resistor,3K,R78,1
Resistor,4.7K,"R1,R2,R3,R4,R5,R6,R7,R31",8
Resistor,10K,"R34,R37,R48,R57,R59,R61,R89,R90,R91,R92,R93,R94",12
Resistor,18K,R43,1
Resistor,24/3216,R82,1
Resistor,30K,"R44,R80",2
Resistor,47K,R29,1
Resistor,56.2K,R47,1
Resistor,270K,R49,1
Resistor,NC,"R63,R65,R67,R69,R75,R77",6
Inductor,NR4018T3R3M,L1,1
Bead,BLM18KG121TN1D,"L2,L3,L4,L5,L6,L7",6
Diode,T1D-CP2R1TY-4T-V1,"D1,D2,D3,D4,D5,D7",6
Diode,1N4148WS,"D8,D9",2
ESD Protection,ESD5Z3.3T1G,"ED1,ED8,ED9",3
ESD Protection,LESD5D5.0CT1G,"ED2,ED3,ED4,ED6,ED7",5
ESD Protection,SMF5.0A,ED5,1
Transistor,KTN2222AS,"Q3,Q4,Q5,Q6,Q7,Q11",6
Transistor,BSS138,Q10,1
Transistor,2N7002,"Q12,Q13",2
IC,MX25L12833FM2I-10G,U1,1
IC,R7FA6M3AH3CFP_100LQFP,U2,1
IC,AP62301WU-7(TSOT26),U3,1
IC,MCP6001T-I/OT,U4,1
IC,GT508L-QN,U5,1
IC,AP2127K-ADJ/AP2127K-3.3TRG1,U6,1
Connector,10031HR-H08G,J1,1
Connector,NC_HEAD_AN_2mm_5P,J4,1
Connector,FH26W-37S-0.3SHW(60),J5,1
Connector,SMW200-H10G,J7,1
Switch,EST-1115,"SW1,SW2",2
Crystal,XRCGB24M000F3M26R0,X1,1
Crystal,SSPT7F-12.5PF20-R,X2,1
Buzzer,TR-5020,BZ1,1
Test Point,TP,"TP1,GP1,TP2,TP3,TP4,TP5,TP6,TP7,TP8,TP10",10
"""

# CSV 데이터를 DataFrame으로 변환
df = pd.read_csv(io.StringIO(csv_data))

# 새 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "BOM"

# 제목 추가
ws['A1'] = "Bill of Materials (BOM)"
ws['A1'].font = Font(bold=True, size=16)
ws.merge_cells('A1:E1')

# 수량 입력 칸 추가
ws['F1'] = "Production Quantity:"
ws['G1'] = 1  # 기본값 설정
ws['F1'].font = Font(bold=True)
ws['G1'].font = Font(bold=True)

# 날짜 추가
today = date.today().strftime("%Y-%m-%d")
ws['A2'] = f"Date: {today}"
ws['A2'].font = Font(italic=True)

# 헤더 추가
headers = ["No", "Item", "Specification", "Package(mm)", "Reference", "Quantity", "Total Quantity", "In Stock", "Required", "Remarks"]
for c, header in enumerate(headers, start=1):
    ws.cell(row=4, column=c, value=header)

# 데이터프레임을 워크시트에 쓰기 (5행부터 시작)
for r, row in enumerate(df.values, start=5):
    ws.cell(row=r, column=1, value=r-4)  # No 열 추가
    for c, value in enumerate(row, start=2):
      if c < 4:  # Package 열 이전까지는 그대로 입력
            ws.cell(row=r, column=c, value=value)
      elif c == 4:  # Package 열에 빈 값 입력
            ws.cell(row=r, column=c, value="")
            ws.cell(row=r, column=c+1, value=value)  # Reference를 한 칸 뒤로
      else:  # 나머지 열들도 한 칸씩 뒤로
            ws.cell(row=r, column=c+1, value=value)
# Total Quantity 계산 공식 추가
for row in range(5, len(df) + 5):
    quantity_cell = ws.cell(row=row, column=6)
    total_quantity_cell = ws.cell(row=row, column=7)
    total_quantity_cell.value = f"={quantity_cell.coordinate}*$G$1"

# Required 계산 공식 추가
for row in range(5, len(df) + 5):
    total_quantity_cell = ws.cell(row=row, column=7)
    in_stock_cell = ws.cell(row=row, column=8)
    required_cell = ws.cell(row=row, column=9)
    required_cell.value = f"=MAX(0, {total_quantity_cell.coordinate}-{in_stock_cell.coordinate})"

# 스타일 설정
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

category_colors = {
    "Resistors": "FFC7CE",
    "Capacitors": "C6EFCE",
    "Semiconductors": "FFEB9C",
    "Connectors": "BDD7EE",
    "LED": "E2EFDA",
    "Others": "D9E1F2"
}

# 스타일 적용
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        if cell.row == 4 or cell.column == 1:  # 헤더 행이거나 No 열인 경우
            cell.font = header_font
            cell.fill = header_fill
        else:
            category = ws.cell(row=cell.row, column=2).value
            if category in category_colors:
                fill_color = category_colors[category]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# 열 너비 조정
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

# Production Quantity 셀에 데이터 유효성 검사 추가
dv = DataValidation(type="whole", operator="greaterThan", formula1=0)
dv.error = "Please enter a positive integer"
dv.errorTitle = "Invalid Input"
ws.add_data_validation(dv)
dv.add('G1')

# 파일 저장
excel_file = 'BOM.xlsx'
wb.save(excel_file)

print(f"Excel 파일이 생성되었습니다: {excel_file}")
