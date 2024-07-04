import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Color

# CSV 데이터를 DataFrame으로 변환
df = pd.read_csv(io.StringIO(csv_data))

# 새 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "BOM"

# 제목 추가
ws['B1'] = "Bill of Materials (BOM)"
ws['B1'].font = Font(bold=True, size=16)
ws.merge_cells('B1:F1')

# 수량 입력 칸 추가
ws['G1'] = "생산수량:"
ws['H1'] = 1  # 기본값 설정
ws['G1'].font = Font(bold=True)
ws['H1'].font = Font(bold=True)
ws['H1'].font = Font(bold=True, color="FF0000")  # 빨간색 설정
ws['H1'].alignment = Alignment(horizontal='center', vertical='center')  # H1 셀 가운데 정렬

# 날짜 추가
today = date.today().strftime("%Y-%m-%d")
ws['B2'] = f"Date: {today}"
ws['B2'].font = Font(italic=True)

# 헤더 추가
headers = ["No", "Item", "Specification", "Package(mm)", "Reference", "Quantity", "Total Qt", "In Stock", "Required", "Remarks"]
for c, header in enumerate(headers, start=2):
    ws.cell(row=4, column=c, value=header)

# 데이터프레임을 워크시트에 쓰기 (5행부터 시작)
for r, row in enumerate(df.values, start=5):
    ws.cell(row=r, column=2, value=r-4)  # No 열 추가
    for c, value in enumerate(row, start=3):
        if c < 5:  # Package 열 이전까지는 그대로 입력
            ws.cell(row=r, column=c, value=value)
        elif c == 5:  # Package 열에 빈 값 입력
            ws.cell(row=r, column=c, value="")
            ws.cell(row=r, column=c+1, value=value)  # Reference를 한 칸 뒤로
        else:  # 나머지 열들도 한 칸씩 뒤로
            ws.cell(row=r, column=c+1, value=value)

# 같은 카테고리 셀 병합
current_category = None
start_row = 5

for row in range(5, ws.max_row + 1):
    category = ws.cell(row=row, column=3).value
    if category != current_category:
        if current_category is not None:
            # 이전 카테고리 셀 병합
            if start_row < row - 1:
                ws.merge_cells(f'C{start_row}:C{row-1}')
        current_category = category
        start_row = row

# 마지막 카테고리 셀 병합
if start_row < ws.max_row:
    ws.merge_cells(f'C{start_row}:C{ws.max_row}')

# 병합된 셀의 정렬 조정
for merged_range in ws.merged_cells.ranges:
    if merged_range.min_col == 3:  # C열(카테고리 열)
        ws.cell(row=merged_range.min_row, column=3).alignment = Alignment(horizontal='center', vertical='center')

# Total Quantity 계산 공식 추가
for row in range(5, len(df) + 5):
    quantity_cell = ws.cell(row=row, column=7)
    total_quantity_cell = ws.cell(row=row, column=8)
    total_quantity_cell.value = f"={quantity_cell.coordinate}*$H$1"

# Required 계산 공식 추가
for row in range(5, len(df) + 5):
    total_quantity_cell = ws.cell(row=row, column=8)
    in_stock_cell = ws.cell(row=row, column=9)
    required_cell = ws.cell(row=row, column=10)
    required_cell.value = f"=MAX(0, {total_quantity_cell.coordinate}-{in_stock_cell.coordinate})"

# 스타일 설정 (변경된 부분)
header_font = Font(bold=True)
header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

category_colors = {
    "Resistors": "FFC7CE",
    "Capacitors": "C6EFCE",
    "Semiconductors": "FFEB9C",
    "Connectors": "BDD7EE",
    "LED": "E2EFDA",
    "Others": "D9E1F2"
}

# 스타일 적용 (변경된 부분)
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        if cell.row == 4 or cell.column == 2:  # 헤더 행이거나 No 열인 경우
            cell.font = header_font
            cell.fill = header_fill
        else:
            category = ws.cell(row=cell.row, column=3).value
            if category in category_colors:
                fill_color = category_colors[category]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# A열의 스타일 제거
for cell in ws['A']:
    cell.border = None
    cell.fill = PatternFill(fill_type=None)

# 외곽 테두리 두껍게 설정 (새로 추가된 부분)
last_row = ws.max_row
last_col = ws.max_column

for col in range(2, last_col + 1):
    ws.cell(row=4, column=col).border = Border(top=Side(style='thick'), left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=last_row, column=col).border = Border(bottom=Side(style='thick'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

for row in range(4, last_row + 1):
    ws.cell(row=row, column=2).border = Border(left=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
    ws.cell(row=row, column=last_col).border = Border(right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))

# 모서리 셀 테두리 수정 (새로 추가된 부분)
ws.cell(row=4, column=2).border = Border(left=Side(style='thick'), top=Side(style='thick'), right=Side(style='thin'), bottom=Side(style='thin'))
ws.cell(row=4, column=last_col).border = Border(right=Side(style='thick'), top=Side(style='thick'), left=Side(style='thin'), bottom=Side(style='thin'))
ws.cell(row=last_row, column=2).border = Border(left=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thin'), top=Side(style='thin'))
ws.cell(row=last_row, column=last_col).border = Border(right=Side(style='thick'), bottom=Side(style='thick'), left=Side(style='thin'), top=Side(style='thin'))

# 열 너비 조정
width = 10  # 기본 너비
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width
    ws.column_dimensions['B'].width = 6  # A열의 너비를 6으로 설정
    ws.column_dimensions['F'].width = 40  # E열의 너비를 40으로 설정
    ws.column_dimensions['G'].width = width  # F열의 너비를 12으로 설정
    ws.column_dimensions['H'].width = width  # E열의 너비를 40으로 설정
    ws.column_dimensions['I'].width = width  # E열의 너비를 40으로 설정
    ws.column_dimensions['J'].width = width  # E열의 너비를 40으로 설정
    ws.column_dimensions['K'].width = 25  # J열의 너비를 16으로 설정

# Production Quantity 셀에 데이터 유효성 검사 추가
dv = DataValidation(type="whole", operator="greaterThan", formula1=0)
dv.error = "Please enter a positive integer"
dv.errorTitle = "Invalid Input"
ws.add_data_validation(dv)
dv.add('H1')

# 파일 저장
excel_file = 'BOM_V3.xlsx'
wb.save(excel_file)

print(f"Excel 파일이 생성되었습니다: {excel_file}")
